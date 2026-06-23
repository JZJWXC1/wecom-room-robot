from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import subprocess
import time
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings
from app.services.feishu import FeishuClient


MAX_IMAGE_HEIGHT = 1800
DEFAULT_ROW_HEIGHT = 28
DEFAULT_COLUMN_WIDTH = 96
CELL_PADDING_X = 8
CELL_PADDING_Y = 5


@dataclass
class RenderedSheet:
    paths: list[Path]
    rows: int
    columns: int


class InventoryImageSyncer:
    def __init__(self, client: FeishuClient | None = None) -> None:
        self.client = client or FeishuClient()

    async def refresh_if_changed(self, *, force: bool = False) -> dict[str, Any]:
        if not settings.feishu_inventory_sheet_token:
            return {"enabled": False, "changed": False, "reason": "missing_sheet_token"}

        state = self._load_state()
        if not force and self._is_recently_checked(state) and self._current_images():
            return {
                "enabled": True,
                "changed": False,
                "reason": "check_interval",
                "signature": state.get("signature", ""),
            }

        sheet_values = await self.client.read_spreadsheet_values(
            spreadsheet_token=settings.feishu_inventory_sheet_token,
        )
        signature = self._sheet_values_signature(sheet_values)
        if (
            not force
            and signature == state.get("signature")
            and self._current_images()
        ):
            state["checked_at"] = time.time()
            self._save_state(state)
            return {
                "enabled": True,
                "changed": False,
                "reason": "unchanged",
                "signature": signature,
                "revision": sheet_values.get("revision"),
            }

        with tempfile.TemporaryDirectory() as directory:
            xlsx_path = Path(directory) / "inventory.xlsx"
            await self.client.export_sheet_xlsx(
                sheet_token=settings.feishu_inventory_sheet_token,
                target_path=xlsx_path,
            )
            rendered = self.render_xlsx_to_inventory_images_from_values(
                xlsx_path,
                sheet_values.get("values") or [],
            )

        self._save_state(
            {
                "signature": signature,
                "checked_at": time.time(),
                "rendered_at": time.time(),
                "image_paths": [str(path) for path in rendered.paths],
                "rows": rendered.rows,
                "columns": rendered.columns,
                "sheet_id": sheet_values.get("sheet_id"),
                "sheet_title": sheet_values.get("title"),
                "range": sheet_values.get("range"),
                "revision": sheet_values.get("revision"),
            }
        )
        self._clear_ocr_cache()
        return {
            "enabled": True,
            "changed": True,
            "signature": signature,
            "image_paths": [str(path) for path in rendered.paths],
            "rows": rendered.rows,
            "columns": rendered.columns,
            "revision": sheet_values.get("revision"),
        }

    def render_values_to_inventory_images(self, values: list[list[Any]]) -> RenderedSheet:
        rows = [
            [str(cell or "").strip() for cell in row]
            for row in values
        ]
        bounds = self._matrix_content_bounds(rows)
        if bounds is None:
            raise ValueError("Inventory sheet has no content")
        min_row, max_row, min_col, max_col = bounds
        matrix = [
            row[min_col : max_col + 1]
            for row in rows[min_row : max_row + 1]
        ]
        return self._render_matrix_to_inventory_images(matrix)

    def render_xlsx_to_inventory_images_from_values(
        self,
        xlsx_path: Path,
        values: list[list[Any]],
    ) -> RenderedSheet:
        libreoffice = shutil.which("libreoffice") or shutil.which("soffice")
        pdftoppm = shutil.which("pdftoppm")
        if not libreoffice or not pdftoppm:
            try:
                return self.render_xlsx_to_inventory_images(xlsx_path)
            except Exception:
                return self.render_values_to_inventory_images(values)

        rows = [[str(cell or "").strip() for cell in row] for row in values]
        bounds = self._matrix_content_bounds(rows)
        if bounds is None:
            raise ValueError("Inventory sheet has no content")
        min_row, max_row, min_col, max_col = bounds

        with tempfile.TemporaryDirectory() as directory:
            work_dir = Path(directory)
            prepared_xlsx = work_dir / "inventory-print-area.xlsx"
            self._write_xlsx_with_print_area(
                source_path=xlsx_path,
                target_path=prepared_xlsx,
                min_row=min_row + 1,
                max_row=max_row + 1,
                min_col=min_col + 1,
                max_col=max_col + 1,
            )
            pdf_dir = work_dir / "pdf"
            pdf_dir.mkdir()
            subprocess.run(
                [
                    libreoffice,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(pdf_dir),
                    str(prepared_xlsx),
                ],
                check=True,
                timeout=180,
                env={**self._subprocess_env(), "HOME": str(work_dir)},
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            pdf_paths = list(pdf_dir.glob("*.pdf"))
            if not pdf_paths:
                raise RuntimeError("LibreOffice did not produce a PDF")

            temp_dir = settings.room_database_path / ".inventory-render-tmp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            self._clear_render_temp(temp_dir)
            output_prefix = temp_dir / "inventory"
            subprocess.run(
                [
                    pdftoppm,
                    "-png",
                    "-r",
                    "160",
                    str(pdf_paths[0]),
                    str(output_prefix),
                ],
                check=True,
                timeout=180,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )

        rendered_paths = sorted(
            temp_dir.glob("inventory-*.png"),
            key=self._rendered_pdf_page_index,
        )
        rendered_paths = self._select_primary_pdf_pages(rendered_paths)
        if not rendered_paths:
            raise RuntimeError("pdftoppm did not produce PNG images")
        cropped_paths = [self._crop_png_whitespace(path) for path in rendered_paths]
        final_paths = self._replace_inventory_images(cropped_paths)
        return RenderedSheet(
            paths=final_paths,
            rows=max_row - min_row + 1,
            columns=max_col - min_col + 1,
        )

    def _sheet_values_signature(self, sheet_values: dict[str, Any]) -> str:
        payload = json.dumps(
            {
                "sheet_id": sheet_values.get("sheet_id"),
                "range": sheet_values.get("range"),
                "revision": sheet_values.get("revision"),
                "values": sheet_values.get("values") or [],
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def render_csv_to_inventory_images(self, csv_path: Path) -> RenderedSheet:
        import csv

        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            return self.render_values_to_inventory_images(
                [
                    [cell.strip() for cell in row]
                    for row in csv.reader(handle)
                ]
            )

    def render_xlsx_to_inventory_images(self, xlsx_path: Path) -> RenderedSheet:
        from PIL import Image, ImageDraw, ImageFont

        workbook = self._load_workbook_for_render(xlsx_path)
        sheet = workbook.active
        bounds = self._content_bounds(sheet)
        if bounds is None:
            raise ValueError("Inventory sheet has no content")
        min_row, max_row, min_col, max_col = bounds

        font = self._load_font(ImageFont, 18)
        bold_font = self._load_font(ImageFont, 18, bold=True)
        column_widths = self._column_widths(sheet, min_col, max_col)
        row_heights = self._row_heights(sheet, min_row, max_row)
        total_width = sum(column_widths) + 1

        temp_dir = settings.room_database_path / ".inventory-render-tmp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        self._clear_render_temp(temp_dir)

        paths: list[Path] = []
        page_start_row = min_row
        page_height = 1
        for row_number in range(min_row, max_row + 1):
            row_height = row_heights[row_number - min_row]
            if page_start_row < row_number and page_height + row_height > MAX_IMAGE_HEIGHT:
                paths.append(
                    self._render_page(
                        Image,
                        ImageDraw,
                        sheet,
                        temp_dir,
                        len(paths) + 1,
                        page_start_row,
                        row_number - 1,
                        min_row,
                        min_col,
                        max_col,
                        column_widths,
                        row_heights,
                        total_width,
                        font,
                        bold_font,
                    )
                )
                page_start_row = row_number
                page_height = 1
            page_height += row_height

        paths.append(
            self._render_page(
                Image,
                ImageDraw,
                sheet,
                temp_dir,
                len(paths) + 1,
                page_start_row,
                max_row,
                min_row,
                min_col,
                max_col,
                column_widths,
                row_heights,
                total_width,
                font,
                bold_font,
            )
        )

        final_paths = self._replace_inventory_images(paths)
        return RenderedSheet(
            paths=final_paths,
            rows=max_row - min_row + 1,
            columns=max_col - min_col + 1,
        )

    def _load_workbook_for_render(self, xlsx_path: Path) -> Any:
        try:
            return load_workbook(xlsx_path, data_only=True)
        except ValueError as exc:
            if "could not read stylesheet" not in str(exc):
                raise
            with tempfile.TemporaryDirectory() as directory:
                sanitized_path = Path(directory) / "inventory-sanitized.xlsx"
                self._write_xlsx_with_sanitized_styles(xlsx_path, sanitized_path)
                return load_workbook(sanitized_path, data_only=True)

    def _write_xlsx_with_sanitized_styles(self, source_path: Path, target_path: Path) -> None:
        with zipfile.ZipFile(source_path, "r") as source, zipfile.ZipFile(
            target_path,
            "w",
            zipfile.ZIP_DEFLATED,
        ) as target:
            for info in source.infolist():
                content = source.read(info.filename)
                if info.filename == "xl/styles.xml":
                    content = self._sanitize_styles_xml(content)
                target.writestr(info, content)

    def _sanitize_styles_xml(self, styles_xml: bytes) -> bytes:
        text = styles_xml.decode("utf-8", errors="replace")

        def replace_rgb(match: re.Match[str]) -> str:
            red, green, blue = (
                max(0, min(255, int(match.group(name))))
                for name in ("red", "green", "blue")
            )
            return f'rgb="FF{red:02X}{green:02X}{blue:02X}"'

        text = re.sub(
            r'rgb="(?:FF)?RGB\(\s*(?P<red>\d+)\s*,\s*(?P<green>\d+)\s*,\s*(?P<blue>\d+)\s*\)"',
            replace_rgb,
            text,
        )
        return text.encode("utf-8")

    def _render_matrix_to_inventory_images(self, matrix: list[list[str]]) -> RenderedSheet:
        from PIL import Image, ImageDraw, ImageFont

        font = self._load_font(ImageFont, 18)
        bold_font = self._load_font(ImageFont, 18, bold=True)
        columns = max(len(row) for row in matrix)
        normalized = [row + [""] * (columns - len(row)) for row in matrix]
        column_widths = self._matrix_column_widths(normalized)
        row_heights = [DEFAULT_ROW_HEIGHT for _ in normalized]
        total_width = sum(column_widths) + 1

        temp_dir = settings.room_database_path / ".inventory-render-tmp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        self._clear_render_temp(temp_dir)

        paths: list[Path] = []
        start_index = 0
        page_height = 1
        for index, row_height in enumerate(row_heights):
            if start_index < index and page_height + row_height > MAX_IMAGE_HEIGHT:
                paths.append(
                    self._render_matrix_page(
                        Image,
                        ImageDraw,
                        temp_dir,
                        len(paths) + 1,
                        normalized[start_index:index],
                        column_widths,
                        total_width,
                        font,
                        bold_font,
                    )
                )
                start_index = index
                page_height = 1
            page_height += row_height

        paths.append(
            self._render_matrix_page(
                Image,
                ImageDraw,
                temp_dir,
                len(paths) + 1,
                normalized[start_index:],
                column_widths,
                total_width,
                font,
                bold_font,
            )
        )
        final_paths = self._replace_inventory_images(paths)
        return RenderedSheet(paths=final_paths, rows=len(normalized), columns=columns)

    def _render_matrix_page(
        self,
        Image: Any,
        ImageDraw: Any,
        output_dir: Path,
        page_number: int,
        rows: list[list[str]],
        column_widths: list[int],
        total_width: int,
        font: Any,
        bold_font: Any,
    ) -> Path:
        total_height = len(rows) * DEFAULT_ROW_HEIGHT + 1
        image = Image.new("RGB", (total_width, total_height), "white")
        draw = ImageDraw.Draw(image)
        y = 0
        for row_index, row in enumerate(rows):
            x = 0
            for col_index, value in enumerate(row):
                width = column_widths[col_index]
                fill = "#f2f6ff" if row_index == 0 else "white"
                draw.rectangle((x, y, x + width, y + DEFAULT_ROW_HEIGHT), fill=fill, outline="#b7b7b7")
                if value:
                    active_font = bold_font if row_index == 0 else font
                    text = "\n".join(
                        self._wrap_text(
                            draw,
                            value,
                            active_font,
                            max(width - CELL_PADDING_X * 2, 20),
                        )[:2]
                    )
                    draw.multiline_text(
                        (x + CELL_PADDING_X, y + CELL_PADDING_Y),
                        text,
                        fill="#111111",
                        font=active_font,
                        spacing=2,
                    )
                x += width
            y += DEFAULT_ROW_HEIGHT
        output_path = output_dir / f"inventory_{page_number:02d}.png"
        image.save(output_path, "PNG")
        return output_path

    def _render_page(
        self,
        Image: Any,
        ImageDraw: Any,
        sheet: Any,
        output_dir: Path,
        page_number: int,
        page_start_row: int,
        page_end_row: int,
        min_row: int,
        min_col: int,
        max_col: int,
        column_widths: list[int],
        row_heights: list[int],
        total_width: int,
        font: Any,
        bold_font: Any,
    ) -> Path:
        start_index = page_start_row - min_row
        page_row_heights = row_heights[start_index : start_index + page_end_row - page_start_row + 1]
        total_height = sum(page_row_heights) + 1
        image = Image.new("RGB", (total_width, total_height), "white")
        draw = ImageDraw.Draw(image)

        column_offsets = self._offsets(column_widths)
        row_offsets = self._offsets(page_row_heights)
        y = 0
        for row_number in range(page_start_row, page_end_row + 1):
            row_height = row_heights[row_number - min_row]
            x = 0
            for col_number in range(min_col, max_col + 1):
                merged_range = self._merged_range_for_cell(sheet, row_number, col_number)
                if merged_range is not None:
                    if (
                        row_number != merged_range.min_row
                        or col_number != merged_range.min_col
                        or merged_range.min_row < page_start_row
                    ):
                        x += column_widths[col_number - min_col]
                        continue
                    draw_min_col = max(merged_range.min_col, min_col)
                    draw_max_col = min(merged_range.max_col, max_col)
                    draw_min_row = max(merged_range.min_row, page_start_row)
                    draw_max_row = min(merged_range.max_row, page_end_row)
                    cell_x = column_offsets[draw_min_col - min_col]
                    cell_y = row_offsets[draw_min_row - page_start_row]
                    merged_width = sum(
                        column_widths[draw_min_col - min_col : draw_max_col - min_col + 1]
                    )
                    merged_height = sum(
                        row_heights[draw_min_row - min_row : draw_max_row - min_row + 1]
                    )
                    self._draw_cell(
                        draw,
                        sheet.cell(row=merged_range.min_row, column=merged_range.min_col),
                        cell_x,
                        cell_y,
                        merged_width,
                        merged_height,
                        font,
                        bold_font,
                    )
                    x += column_widths[col_number - min_col]
                    continue
                cell = sheet.cell(row=row_number, column=col_number)
                col_width = column_widths[col_number - min_col]
                self._draw_cell(
                    draw,
                    cell,
                    x,
                    y,
                    col_width,
                    row_height,
                    font,
                    bold_font,
                )
                x += col_width
            y += row_height

        output_path = output_dir / f"inventory_{page_number:02d}.png"
        image.save(output_path, "PNG")
        return output_path

    def _merged_range_for_cell(self, sheet: Any, row_number: int, col_number: int) -> Any | None:
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row_number <= merged_range.max_row
                and merged_range.min_col <= col_number <= merged_range.max_col
            ):
                return merged_range
        return None

    def _offsets(self, sizes: list[int]) -> list[int]:
        offsets: list[int] = []
        position = 0
        for size in sizes:
            offsets.append(position)
            position += size
        return offsets

    def _draw_cell(
        self,
        draw: Any,
        cell: Cell,
        x: int,
        y: int,
        width: int,
        height: int,
        font: Any,
        bold_font: Any,
    ) -> None:
        fill = self._fill_color(cell.fill)
        draw.rectangle((x, y, x + width, y + height), fill=fill, outline="#b7b7b7")
        value = self._cell_text(cell)
        if not value:
            return
        active_font = bold_font if cell.font and cell.font.bold else font
        text_color = self._font_color(cell)
        lines = self._wrap_text(draw, value, active_font, max(width - CELL_PADDING_X * 2, 20))
        self._draw_aligned_text(
            draw,
            lines[:4],
            x,
            y,
            width,
            height,
            active_font,
            text_color,
            horizontal=str(cell.alignment.horizontal or ""),
            vertical=str(cell.alignment.vertical or ""),
        )

    def _draw_aligned_text(
        self,
        draw: Any,
        lines: list[str],
        x: int,
        y: int,
        width: int,
        height: int,
        font: Any,
        fill: str,
        *,
        horizontal: str = "",
        vertical: str = "",
    ) -> None:
        if not lines:
            return
        line_height = self._text_height(draw, font)
        spacing = 2
        text_height = len(lines) * line_height + max(0, len(lines) - 1) * spacing
        if vertical in {"center", "middle", "distributed"}:
            text_y = y + max(CELL_PADDING_Y, (height - text_height) // 2)
        elif vertical == "bottom":
            text_y = y + max(CELL_PADDING_Y, height - text_height - CELL_PADDING_Y)
        else:
            text_y = y + CELL_PADDING_Y
        for line in lines:
            line_width = draw.textlength(line, font=font)
            if horizontal in {"center", "centerContinuous", "distributed"}:
                text_x = x + max(CELL_PADDING_X, (width - int(line_width)) // 2)
            elif horizontal == "right":
                text_x = x + max(CELL_PADDING_X, width - int(line_width) - CELL_PADDING_X)
            else:
                text_x = x + CELL_PADDING_X
            draw.text((text_x, text_y), line, fill=fill, font=font)
            text_y += line_height + spacing

    def _text_height(self, draw: Any, font: Any) -> int:
        bbox = draw.textbbox((0, 0), "国", font=font)
        return max(1, bbox[3] - bbox[1])

    def _content_bounds(self, sheet: Any) -> tuple[int, int, int, int] | None:
        header_bounds = self._inventory_header_bounds(sheet)
        if header_bounds is not None:
            return header_bounds

        min_row = min_col = 10**9
        max_row = max_col = 0
        for row in sheet.iter_rows():
            for cell in row:
                if self._cell_text(cell):
                    min_row = min(min_row, cell.row)
                    max_row = max(max_row, cell.row)
                    min_col = min(min_col, cell.column)
                    max_col = max(max_col, cell.column)
        if max_row == 0:
            return None
        return min_row, max_row, min_col, max_col

    def _inventory_header_bounds(self, sheet: Any) -> tuple[int, int, int, int] | None:
        for row_number in range(1, sheet.max_row + 1):
            cells = [
                self._cell_text(sheet.cell(row=row_number, column=column_number))
                for column_number in range(1, min(sheet.max_column, 50) + 1)
            ]
            if "小区" not in cells or "房号" not in cells:
                continue
            header_columns = [
                index
                for index, value in enumerate(cells, start=1)
                if value.strip()
            ]
            if not header_columns:
                continue
            min_col = min(header_columns)
            max_col = max(header_columns)
            min_row = 10**9
            max_row = 0
            for data_row in range(1, sheet.max_row + 1):
                has_content = any(
                    self._cell_text(sheet.cell(row=data_row, column=column_number))
                    for column_number in range(min_col, max_col + 1)
                )
                if has_content:
                    min_row = min(min_row, data_row)
                    max_row = max(max_row, data_row)
            if max_row:
                return min_row, max_row, min_col, max_col
        return None

    def _column_widths(self, sheet: Any, min_col: int, max_col: int) -> list[int]:
        widths: list[int] = []
        for col_number in range(min_col, max_col + 1):
            letter = get_column_letter(col_number)
            width = sheet.column_dimensions[letter].width
            pixels = int((width or 12) * 7 + 12)
            widths.append(max(60, min(260, pixels or DEFAULT_COLUMN_WIDTH)))
        return widths

    def _row_heights(self, sheet: Any, min_row: int, max_row: int) -> list[int]:
        heights: list[int] = []
        for row_number in range(min_row, max_row + 1):
            height = sheet.row_dimensions[row_number].height
            pixels = int((height or 21) * 1.33)
            heights.append(max(DEFAULT_ROW_HEIGHT, min(96, pixels)))
        return heights

    def _write_xlsx_with_print_area(
        self,
        *,
        source_path: Path,
        target_path: Path,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> None:
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_namespace = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        package_rel_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
        ET.register_namespace("", namespace)
        workbook_name = "xl/workbook.xml"
        workbook_rels_name = "xl/_rels/workbook.xml.rels"
        with zipfile.ZipFile(source_path, "r") as source:
            workbook_xml = source.read(workbook_name)
            root = ET.fromstring(workbook_xml)
            sheets = root.find(f"{{{namespace}}}sheets")
            if sheets is None or not list(sheets):
                raise RuntimeError("XLSX workbook has no sheets")
            first_sheet = list(sheets)[0]
            sheet_name = str(first_sheet.attrib.get("name") or "Sheet1")
            relationship_id = first_sheet.attrib.get(f"{{{rel_namespace}}}id")
            if not relationship_id:
                raise RuntimeError("XLSX first sheet relationship id is empty")
            rels_root = ET.fromstring(source.read(workbook_rels_name))
            worksheet_name = ""
            for relationship in rels_root.findall(f"{{{package_rel_namespace}}}Relationship"):
                if relationship.attrib.get("Id") == relationship_id:
                    target = str(relationship.attrib.get("Target") or "")
                    worksheet_name = str(PurePosixPath("xl") / target)
                    break
            if not worksheet_name:
                raise RuntimeError("XLSX first worksheet path is empty")
            defined_names = root.find(f"{{{namespace}}}definedNames")
            if defined_names is None:
                defined_names = ET.Element(f"{{{namespace}}}definedNames")
                insert_index = len(root)
                for index, child in enumerate(list(root)):
                    if child.tag.endswith("calcPr"):
                        insert_index = index
                        break
                root.insert(insert_index, defined_names)
            for item in list(defined_names):
                if (
                    item.attrib.get("name") == "_xlnm.Print_Area"
                    and item.attrib.get("localSheetId", "0") == "0"
                ):
                    defined_names.remove(item)
            min_col_letter = get_column_letter(min_col)
            max_col_letter = get_column_letter(max_col)
            escaped_name = sheet_name.replace("'", "''")
            print_area = ET.SubElement(
                defined_names,
                f"{{{namespace}}}definedName",
                {"name": "_xlnm.Print_Area", "localSheetId": "0"},
            )
            print_area.text = (
                f"'{escaped_name}'!"
                f"${min_col_letter}${min_row}:${max_col_letter}${max_row}"
            )
            patched_workbook = ET.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )
            with zipfile.ZipFile(target_path, "w", zipfile.ZIP_DEFLATED) as target:
                for info in source.infolist():
                    if info.filename == workbook_name:
                        target.writestr(info, patched_workbook)
                    elif info.filename == worksheet_name:
                        target.writestr(
                            info,
                            self._patch_worksheet_print_settings(
                                source.read(info.filename),
                                namespace,
                            ),
                        )
                    else:
                        target.writestr(info, source.read(info.filename))

    def _patch_worksheet_print_settings(
        self,
        worksheet_xml: bytes,
        namespace: str,
    ) -> bytes:
        root = ET.fromstring(worksheet_xml)
        for child_name in ("rowBreaks", "colBreaks", "pageMargins", "pageSetup", "printOptions"):
            for item in list(root.findall(f"{{{namespace}}}{child_name}")):
                root.remove(item)
        sheet_pr = root.find(f"{{{namespace}}}sheetPr")
        if sheet_pr is None:
            sheet_pr = ET.Element(f"{{{namespace}}}sheetPr")
            root.insert(0, sheet_pr)
        page_setup_pr = sheet_pr.find(f"{{{namespace}}}pageSetUpPr")
        if page_setup_pr is None:
            page_setup_pr = ET.SubElement(sheet_pr, f"{{{namespace}}}pageSetUpPr")
        page_setup_pr.attrib["fitToPage"] = "1"

        ET.SubElement(
            root,
            f"{{{namespace}}}printOptions",
            {"horizontalCentered": "1"},
        )
        ET.SubElement(
            root,
            f"{{{namespace}}}pageMargins",
            {
                "left": "0.1",
                "right": "0.1",
                "top": "0.1",
                "bottom": "0.1",
                "header": "0",
                "footer": "0",
            },
        )
        ET.SubElement(
            root,
            f"{{{namespace}}}pageSetup",
            {
                "paperSize": "9",
                "orientation": "landscape",
                "fitToWidth": "1",
                "fitToHeight": "0",
            },
        )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def _crop_png_whitespace(self, path: Path) -> Path:
        from PIL import Image, ImageChops

        with Image.open(path) as image:
            rgb = image.convert("RGB")
            background = Image.new("RGB", rgb.size, "white")
            diff = ImageChops.difference(rgb, background)
            mask = diff.convert("L").point(lambda value: 255 if value > 8 else 0)
            bbox = mask.getbbox()
            if not bbox:
                return path
            margin = 10
            left = max(0, bbox[0] - margin)
            top = max(0, bbox[1] - margin)
            right = min(rgb.width, bbox[2] + margin)
            bottom = min(rgb.height, bbox[3] + margin)
            rgb.crop((left, top, right, bottom)).save(path, "PNG")
        return path

    def _subprocess_env(self) -> dict[str, str]:
        env = dict(os.environ)
        env.setdefault("LANG", "C.UTF-8")
        env.setdefault("LC_ALL", "C.UTF-8")
        return env

    def _matrix_content_bounds(
        self,
        rows: list[list[str]],
    ) -> tuple[int, int, int, int] | None:
        min_row = min_col = 10**9
        max_row = max_col = -1
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                if value.strip():
                    min_row = min(min_row, row_index)
                    max_row = max(max_row, row_index)
                    min_col = min(min_col, col_index)
                    max_col = max(max_col, col_index)
        if max_row < 0:
            return None
        return min_row, max_row, min_col, max_col

    def _matrix_column_widths(self, rows: list[list[str]]) -> list[int]:
        columns = max(len(row) for row in rows)
        widths: list[int] = []
        for col_index in range(columns):
            max_chars = max(
                (len(row[col_index]) if col_index < len(row) else 0)
                for row in rows
            )
            widths.append(max(72, min(260, max_chars * 18 + 28)))
        return widths

    def _replace_inventory_images(self, rendered_paths: list[Path]) -> list[Path]:
        settings.room_database_path.mkdir(parents=True, exist_ok=True)
        final_paths = [
            settings.room_database_path / f"inventory_{index:02d}.png"
            for index, _ in enumerate(rendered_paths, start=1)
        ]
        for old_path in self._current_images():
            old_path.unlink()
        for rendered_path, final_path in zip(rendered_paths, final_paths):
            rendered_path.replace(final_path)
        try:
            rendered_paths[0].parent.rmdir()
        except OSError:
            pass
        return final_paths

    def _clear_render_temp(self, temp_dir: Path) -> None:
        for pattern in ("inventory_*.png", "inventory-*.png"):
            for old_path in temp_dir.glob(pattern):
                old_path.unlink()

    def _rendered_pdf_page_index(self, path: Path) -> int:
        match = re.search(r"-(\d+)\.png$", path.name)
        return int(match.group(1)) if match else 0

    def _select_primary_pdf_pages(self, paths: list[Path]) -> list[Path]:
        if not paths:
            return []
        from PIL import Image

        selected: list[Path] = []
        with Image.open(paths[0]) as first_image:
            expected_width = first_image.width
        for path in paths:
            with Image.open(path) as image:
                width = image.width
            if selected and width < expected_width * 0.9:
                break
            selected.append(path)
        return selected

    def _current_images(self) -> list[Path]:
        paths = sorted(settings.room_database_path.parent.glob(settings.inventory_image_glob))
        if not paths and settings.inventory_image_path.exists():
            paths = [settings.inventory_image_path]
        return [path for path in paths if path.is_file()]

    def _load_state(self) -> dict[str, Any]:
        path = settings.inventory_image_sync_state_path
        if not path.exists():
            return {}
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        return data if isinstance(data, dict) else {}

    def _save_state(self, state: dict[str, Any]) -> None:
        path = settings.inventory_image_sync_state_path
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(f"{path.suffix}.tmp")
        tmp_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_path.replace(path)

    def _is_recently_checked(self, state: dict[str, Any]) -> bool:
        checked_at = float(state.get("checked_at") or 0)
        return time.time() - checked_at < settings.feishu_inventory_sheet_check_seconds

    def _file_sha256(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _clear_ocr_cache(self) -> None:
        try:
            settings.inventory_image_cache_path.unlink()
        except FileNotFoundError:
            pass

    def _cell_text(self, cell: Cell) -> str:
        value = cell.value
        if value is None:
            return ""
        return str(value).strip()

    def _fill_color(self, fill: PatternFill) -> str:
        if not self._has_visible_fill(fill):
            return "white"
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            return f"#{color.rgb[-6:]}"
        if color.type == "indexed" and color.indexed:
            indexed = {
                5: "#ffff00",
                6: "#ff0000",
                10: "#00ff00",
                13: "#ff00ff",
                22: "#c0c0c0",
            }
            return indexed.get(color.indexed, "white")
        return "white"

    def _has_visible_fill(self, fill: PatternFill) -> bool:
        return bool(fill and fill.fill_type and fill.fgColor and fill.fgColor.type != "auto")

    def _font_color(self, cell: Cell) -> str:
        color = cell.font.color if cell.font else None
        if color and color.type == "rgb" and color.rgb:
            return f"#{color.rgb[-6:]}"
        return "#111111"

    def _wrap_text(self, draw: Any, text: str, font: Any, width: int) -> list[str]:
        lines: list[str] = []
        for paragraph in text.splitlines() or [""]:
            current = ""
            for char in paragraph:
                candidate = current + char
                if draw.textlength(candidate, font=font) <= width or not current:
                    current = candidate
                else:
                    lines.append(current)
                    current = char
            lines.append(current)
        return lines

    def _load_font(self, ImageFont: Any, size: int, *, bold: bool = False) -> Any:
        candidates = [
            "C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"
            if bold
            else "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for candidate in candidates:
            path = Path(candidate)
            if path.exists():
                return ImageFont.truetype(str(path), size=size)
        return ImageFont.load_default()
