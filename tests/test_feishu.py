import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import httpx
from openpyxl import Workbook

from app.config import settings
from app.services.feishu import FeishuClient
from app.services.inventory_image_sync import InventoryImageSyncer


class FeishuClientTests(unittest.IsolatedAsyncioTestCase):
    async def test_reads_bitable_records_as_dataframe(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_app_token = settings.feishu_bitable_app_token
        previous_table_id = settings.feishu_bitable_table_id
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_bitable_app_token = "app_token"
            settings.feishu_bitable_table_id = "table_id"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                self.assertTrue(
                    request.url.path.endswith(
                        "/bitable/v1/apps/app_token/tables/table_id/records/search"
                    )
                )
                return httpx.Response(
                    200,
                    json={
                        "code": 0,
                        "data": {
                            "items": [
                                {
                                    "record_id": "rec_1",
                                    "fields": {
                                        "小区": "小洋坝",
                                        "租金": 1800,
                                        "标签": [{"text": "可短租"}],
                                    },
                                }
                            ],
                            "has_more": False,
                        },
                    },
                )

            client = FeishuClient(transport=httpx.MockTransport(handler))
            frame = await client.read_bitable_dataframe()

            self.assertEqual(len(frame), 1)
            self.assertEqual(frame.iloc[0]["小区"], "小洋坝")
            self.assertEqual(frame.iloc[0]["租金"], "1800")
            self.assertEqual(frame.iloc[0]["标签"], "可短租")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_bitable_app_token = previous_app_token
            settings.feishu_bitable_table_id = previous_table_id

    async def test_creates_bitable_record(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                self.assertTrue(
                    request.url.path.endswith(
                        "/bitable/v1/apps/app_token/tables/table_id/records"
                    )
                )
                self.assertIn("待人工跟进", request.content.decode("utf-8"))
                return httpx.Response(
                    200,
                    json={
                        "code": 0,
                        "data": {
                            "record": {
                                "record_id": "rec_1",
                                "fields": {"跟进状态": "待人工跟进"},
                            }
                        },
                    },
                )

            client = FeishuClient(transport=httpx.MockTransport(handler))
            record = await client.create_bitable_record(
                app_token="app_token",
                table_id="table_id",
                fields={"跟进状态": "待人工跟进"},
            )

            self.assertEqual(record["record_id"], "rec_1")
            self.assertEqual(record["fields"]["跟进状态"], "待人工跟进")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret

    async def test_writes_spreadsheet_values(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={
                            "code": 0,
                            "data": {
                                "sheets": [
                                    {"sheet_id": "sheet_1", "title": "总", "index": 0}
                                ]
                            },
                        },
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/values"):
                    body = request.content.decode("utf-8")
                    self.assertEqual(request.method, "PUT")
                    self.assertIn("sheet_1!B2:C3", body)
                    self.assertIn("棠润府", body)
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"revision": 3}},
                    )
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.write_spreadsheet_values(
                start_cell="B2",
                values=[["棠润府", "1-602A"], ["华丰人家", "8-603"]],
            )

            self.assertEqual(result["revision"], 3)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_creates_drive_folder(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                self.assertTrue(request.url.path.endswith("/drive/v1/files/create_folder"))
                body = request.content.decode("utf-8")
                self.assertIn("root_folder", body)
                self.assertIn("棠润府1-602A", body)
                return httpx.Response(
                    200,
                    json={"code": 0, "data": {"token": "folder_1"}},
                )

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.create_drive_folder(
                parent_folder_token="root_folder",
                name="棠润府1-602A",
            )

            self.assertEqual(result["token"], "folder_1")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret

    async def test_sync_drive_media_for_rooms_downloads_only_matching_room_and_kind(self) -> None:
        previous_root = settings.feishu_drive_root_folder_token
        previous_inventory_root = settings.feishu_inventory_drive_folder_token
        settings.feishu_drive_root_folder_token = "root"
        settings.feishu_inventory_drive_folder_token = ""

        class FakeClient(FeishuClient):
            async def list_folder_files(self, folder_token: str) -> list[dict]:
                if folder_token == "root":
                    return [{"name": "拱墅万达 北部软件园", "type": "folder", "token": "area"}]
                if folder_token == "area":
                    return [
                        {"name": "棠润府1-602A", "type": "folder", "token": "room"},
                        {"name": "棠润府10-1004C", "type": "folder", "token": "other-room"},
                    ]
                if folder_token == "room":
                    return [
                        {"name": "a.mp4", "type": "file", "token": "video-token"},
                        {"name": "a.jpg", "type": "file", "token": "image-token"},
                    ]
                if folder_token == "other-room":
                    return [{"name": "other.mp4", "type": "file", "token": "other-token"}]
                return []

            async def download_file(self, file_token: str, target_path: Path) -> Path:
                target_path.parent.mkdir(parents=True, exist_ok=True)
                target_path.write_bytes(file_token.encode("utf-8"))
                return target_path

        try:
            with tempfile.TemporaryDirectory() as directory:
                target_root = Path(directory)
                result = await FakeClient().sync_drive_media_for_rooms(
                    [{"小区": "棠润府", "房号": "1-602A"}],
                    media_kind="video",
                    target_root=target_root,
                )

                downloaded = [Path(path) for path in result["downloaded"]]
                self.assertEqual(len(downloaded), 1)
                self.assertEqual(downloaded[0].name, "a.mp4")
                self.assertTrue(downloaded[0].is_file())
                self.assertFalse((target_root / "images").exists())
                self.assertEqual(result["missing"], [])
        finally:
            settings.feishu_drive_root_folder_token = previous_root
            settings.feishu_inventory_drive_folder_token = previous_inventory_root

    async def test_resolves_inventory_media_folder_from_inventory_drive_root(self) -> None:
        previous_root = settings.feishu_drive_root_folder_token
        previous_inventory_root = settings.feishu_inventory_drive_folder_token
        settings.feishu_drive_root_folder_token = "wrong-root"
        settings.feishu_inventory_drive_folder_token = "inventory-root"

        class FakeClient(FeishuClient):
            async def list_folder_files(self, folder_token: str) -> list[dict]:
                self.seen_folder_token = folder_token
                return [
                    {"name": "记忆文件c", "type": "folder", "token": "memory"},
                    {"name": "房源素材", "type": "folder", "token": "media-root"},
                    {"name": "寓你住一起房源表", "type": "sheet", "token": "sheet-token"},
                ]

        try:
            client = FakeClient()
            token = await client.resolve_inventory_media_folder_token()

            self.assertEqual(token, "media-root")
            self.assertEqual(client.seen_folder_token, "inventory-root")
        finally:
            settings.feishu_drive_root_folder_token = previous_root
            settings.feishu_inventory_drive_folder_token = previous_inventory_root

    async def test_unmerges_spreadsheet_cells(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={
                            "code": 0,
                            "data": {
                                "sheets": [
                                    {"sheet_id": "sheet_1", "title": "总", "index": 0}
                                ]
                            },
                        },
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/unmerge_cells"):
                    self.assertIn("sheet_1!A1:I60", request.content.decode("utf-8"))
                    return httpx.Response(200, json={"code": 0, "data": {}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            await client.unmerge_spreadsheet_cells(range_name="A1:I60")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_inserts_spreadsheet_rows_with_inherited_style(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(200, json={"code": 0, "tenant_access_token": "tenant_token"})
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"sheets": [{"sheet_id": "sheet_1", "title": "总", "index": 0}]}},
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/insert_dimension_range"):
                    self.assertEqual(request.method, "POST")
                    body = request.content.decode("utf-8")
                    self.assertIn('"sheetId":"sheet_1"', body)
                    self.assertIn('"majorDimension":"ROWS"', body)
                    self.assertIn('"startIndex":3', body)
                    self.assertIn('"endIndex":5', body)
                    self.assertIn('"inheritStyle":"BEFORE"', body)
                    return httpx.Response(200, json={"code": 0, "data": {"revision": 4}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.insert_spreadsheet_rows(start_row=4, count=2, inherit_style="BEFORE")

            self.assertEqual(result["revision"], 4)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_deletes_spreadsheet_rows(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(200, json={"code": 0, "tenant_access_token": "tenant_token"})
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"sheets": [{"sheet_id": "sheet_1", "title": "总", "index": 0}]}},
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/dimension_range"):
                    self.assertEqual(request.method, "DELETE")
                    body = request.content.decode("utf-8")
                    self.assertIn('"sheetId":"sheet_1"', body)
                    self.assertIn('"majorDimension":"ROWS"', body)
                    self.assertIn('"startIndex":9', body)
                    self.assertIn('"endIndex":12', body)
                    return httpx.Response(200, json={"code": 0, "data": {"revision": 6}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.delete_spreadsheet_rows(start_row=10, count=3)

            self.assertEqual(result["revision"], 6)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_updates_spreadsheet_row_height(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(200, json={"code": 0, "tenant_access_token": "tenant_token"})
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"sheets": [{"sheet_id": "sheet_1", "title": "总", "index": 0}]}},
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/dimension_range"):
                    self.assertEqual(request.method, "PUT")
                    body = request.content.decode("utf-8")
                    self.assertIn('"sheetId":"sheet_1"', body)
                    self.assertIn('"majorDimension":"ROWS"', body)
                    self.assertIn('"startIndex":3', body)
                    self.assertIn('"endIndex":6', body)
                    self.assertIn('"fixedSize":72', body)
                    return httpx.Response(200, json={"code": 0, "data": {"revision": 5}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.update_spreadsheet_row_height(
                start_row=3,
                end_row=6,
                height_px=72,
            )

            self.assertEqual(result["revision"], 5)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_merges_spreadsheet_cells(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(200, json={"code": 0, "tenant_access_token": "tenant_token"})
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"sheets": [{"sheet_id": "sheet_1", "title": "总", "index": 0}]}},
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/merge_cells"):
                    self.assertEqual(request.method, "POST")
                    body = request.content.decode("utf-8")
                    self.assertIn("sheet_1!A6:A20", body)
                    self.assertIn('"mergeType":"MERGE_ALL"', body)
                    return httpx.Response(200, json={"code": 0, "data": {"revision": 7}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.merge_spreadsheet_cells(range_name="A6:A20")

            self.assertEqual(result["revision"], 7)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_batch_updates_spreadsheet_styles(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(200, json={"code": 0, "tenant_access_token": "tenant_token"})
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/sheets/v3/spreadsheets/sheet_token/sheets/query"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "data": {"sheets": [{"sheet_id": "sheet_1", "title": "总", "index": 0}]}},
                    )
                if request.url.path.endswith("/sheets/v2/spreadsheets/sheet_token/styles_batch_update"):
                    self.assertEqual(request.method, "PUT")
                    body = request.content.decode("utf-8")
                    self.assertIn("sheet_1!B3:I4", body)
                    self.assertIn('"backColor":"#F8CBAD"', body)
                    self.assertIn('"clean":false', body)
                    return httpx.Response(200, json={"code": 0, "data": {"revision": 5}})
                raise AssertionError(f"Unexpected request: {request.url}")

            client = FeishuClient(transport=httpx.MockTransport(handler))
            result = await client.batch_update_spreadsheet_styles(
                updates=[{"ranges": ["B3:I4"], "style": {"backColor": "#F8CBAD"}}]
            )

            self.assertEqual(result["revision"], 5)
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_uploads_drive_file(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                self.assertTrue(request.url.path.endswith("/drive/v1/files/upload_all"))
                body = request.content
                self.assertIn(b"parent_folder", body)
                self.assertIn(b"room.jpg", body)
                self.assertIn(b"image-bytes", body)
                return httpx.Response(
                    200,
                    json={"code": 0, "data": {"file_token": "file_1"}},
                )

            with tempfile.TemporaryDirectory() as directory:
                file_path = Path(directory) / "room.jpg"
                file_path.write_bytes(b"image-bytes")
                client = FeishuClient(transport=httpx.MockTransport(handler))
                result = await client.upload_drive_file(
                    parent_folder_token="parent_folder",
                    file_path=file_path,
                )

            self.assertEqual(result["file_token"], "file_1")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret

    def test_extracts_attachments_recursively(self) -> None:
        client = FeishuClient()
        record = {
            "fields": {
                "房源笔记": [
                    {
                        "children": [
                            {"name": "客厅.jpg", "file_token": "image_token"},
                            {"name": "视频.mp4", "token": "video_token"},
                        ]
                    },
                    {
                        "type": "mention",
                        "mentionType": "Docx",
                        "text": "棠润府1-602A",
                        "token": "doc_token",
                    }
                ],
                "普通链接": {"url": "https://example.com"},
            }
        }

        attachments = client.extract_attachments(record)

        self.assertEqual(
            [item.get("name") for item in attachments],
            ["客厅.jpg", "视频.mp4"],
        )

    async def test_retries_when_tenant_access_token_is_invalid(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"
            auth_tokens = ["old_token", "new_token"]
            auth_calls = 0
            query_tokens: list[str] = []

            def handler(request: httpx.Request) -> httpx.Response:
                nonlocal auth_calls
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    token = auth_tokens[auth_calls]
                    auth_calls += 1
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": token},
                    )
                query_tokens.append(request.headers["Authorization"])
                if query_tokens[-1] == "Bearer old_token":
                    return httpx.Response(
                        400,
                        json={
                            "code": 99991663,
                            "msg": "Invalid access token for authorization.",
                        },
                    )
                return httpx.Response(
                    200,
                    json={
                        "code": 0,
                        "data": {
                            "sheets": [
                                {"sheet_id": "sheet_1", "title": "总", "index": 0}
                            ]
                        },
                    },
                )

            client = FeishuClient(transport=httpx.MockTransport(handler))
            sheets = await client.list_spreadsheet_sheets()

            self.assertEqual(sheets[0]["sheet_id"], "sheet_1")
            self.assertEqual(auth_calls, 2)
            self.assertEqual(query_tokens, ["Bearer old_token", "Bearer new_token"])
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_syncs_drive_media_to_room_database(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                if request.url.path.endswith("/drive/v1/files"):
                    folder_token = request.url.params.get("folder_token")
                    if folder_token == "root":
                        return httpx.Response(
                            200,
                            json={
                                "code": 0,
                                "data": {
                                    "files": [
                                        {
                                            "name": "小洋坝三区12-1003-2",
                                            "type": "folder",
                                            "token": "folder_1",
                                        }
                                    ],
                                    "has_more": False,
                                },
                            },
                        )
                    if folder_token == "folder_1":
                        return httpx.Response(
                            200,
                            json={
                                "code": 0,
                                "data": {
                                    "files": [
                                        {
                                            "name": "看房视频.mp4",
                                            "type": "file",
                                            "token": "video_token",
                                        },
                                        {
                                            "name": "客厅.jpg",
                                            "type": "file",
                                            "token": "image_token",
                                        },
                                        {
                                            "name": "说明.txt",
                                            "type": "file",
                                            "token": "text_token",
                                        },
                                    ],
                                    "has_more": False,
                                },
                            },
                        )
                if request.url.path.endswith("/drive/v1/files/video_token/download"):
                    return httpx.Response(200, content=b"video")
                if request.url.path.endswith("/drive/v1/files/image_token/download"):
                    return httpx.Response(200, content=b"image")
                raise AssertionError(f"Unexpected request: {request.url}")

            with tempfile.TemporaryDirectory() as directory:
                target_root = Path(directory) / "room_database"
                client = FeishuClient(transport=httpx.MockTransport(handler))
                result = await client.sync_drive_media(
                    folder_token="root",
                    target_root=target_root,
                )

                self.assertEqual(len(result["downloaded"]), 2)
                self.assertEqual(result["skipped"], ["说明.txt"])
                self.assertEqual(
                    (target_root / "video" / "小洋坝三区12-1003-2" / "看房视频.mp4").read_bytes(),
                    b"video",
                )
                self.assertEqual(
                    (target_root / "images" / "小洋坝三区12-1003-2" / "客厅.jpg").read_bytes(),
                    b"image",
                )
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret

    async def test_sync_drive_media_skips_existing_local_file(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_state_path = settings.feishu_media_sync_state_path
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                if request.url.path.endswith("/drive/v1/files"):
                    return httpx.Response(
                        200,
                        json={
                            "code": 0,
                            "data": {
                                "files": [
                                    {
                                        "name": "room.jpg",
                                        "type": "file",
                                        "token": "image_token",
                                    }
                                ],
                                "has_more": False,
                            },
                        },
                    )
                raise AssertionError(f"Unexpected request: {request.url}")

            with tempfile.TemporaryDirectory() as directory:
                settings.feishu_media_sync_state_path = Path(directory) / "sync_state.json"
                target_root = Path(directory) / "room_database"
                existing_path = target_root / "images" / "room.jpg"
                existing_path.parent.mkdir(parents=True, exist_ok=True)
                existing_path.write_bytes(b"old-image")

                client = FeishuClient(transport=httpx.MockTransport(handler))
                result = await client.sync_drive_media(
                    folder_token="root",
                    target_root=target_root,
                )

                self.assertEqual(result["downloaded"], [])
                self.assertEqual(result["skipped"], [str(existing_path)])
                self.assertEqual(existing_path.read_bytes(), b"old-image")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_media_sync_state_path = previous_state_path

    async def test_syncs_bitable_attachments_to_room_database(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_app_token = settings.feishu_bitable_app_token
        previous_table_id = settings.feishu_bitable_table_id
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_bitable_app_token = "app_token"
            settings.feishu_bitable_table_id = "table_id"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith(
                    "/bitable/v1/apps/app_token/tables/table_id/records/search"
                ):
                    return httpx.Response(
                        200,
                        json={
                            "code": 0,
                            "data": {
                                "items": [
                                    {
                                        "record_id": "rec_1",
                                        "fields": {
                                            "小区": "棠润府",
                                            "房号": "1-602A",
                                            "房源图片": [
                                                {
                                                    "name": "客厅.jpg",
                                                    "file_token": "image_token",
                                                    "url": "/attachment/image",
                                                }
                                            ],
                                            "房源视频": [
                                                {
                                                    "name": "看房视频.mp4",
                                                    "file_token": "video_token",
                                                    "url": "/attachment/video",
                                                }
                                            ],
                                            "说明": [
                                                {
                                                    "name": "说明.txt",
                                                    "file_token": "text_token",
                                                    "url": "/attachment/text",
                                                }
                                            ],
                                        },
                                    }
                                ],
                                "has_more": False,
                            },
                        },
                    )
                if request.url.path.endswith("/attachment/image"):
                    return httpx.Response(200, content=b"image")
                if request.url.path.endswith("/attachment/video"):
                    return httpx.Response(200, content=b"video")
                raise AssertionError(f"Unexpected request: {request.url}")

            with tempfile.TemporaryDirectory() as directory:
                target_root = Path(directory) / "room_database"
                client = FeishuClient(transport=httpx.MockTransport(handler))
                result = await client.sync_bitable_media(target_root=target_root)

                self.assertEqual(len(result["downloaded"]), 2)
                self.assertEqual(result["skipped"], ["说明.txt"])
                self.assertEqual(
                    (target_root / "images" / "棠润府-1-602A" / "客厅.jpg").read_bytes(),
                    b"image",
                )
                self.assertEqual(
                    (target_root / "video" / "棠润府-1-602A" / "看房视频.mp4").read_bytes(),
                    b"video",
                )
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_bitable_app_token = previous_app_token
            settings.feishu_bitable_table_id = previous_table_id

    async def test_sync_bitable_media_skips_deleted_note_error(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_app_token = settings.feishu_bitable_app_token
        previous_table_id = settings.feishu_bitable_table_id
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_bitable_app_token = "app_token"
            settings.feishu_bitable_table_id = "table_id"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith(
                    "/bitable/v1/apps/app_token/tables/table_id/records/search"
                ):
                    return httpx.Response(
                        200,
                        json={"code": 1002, "msg": "note has been deleted", "data": {}},
                    )
                raise AssertionError(f"Unexpected request: {request.url}")

            with tempfile.TemporaryDirectory() as directory:
                client = FeishuClient(transport=httpx.MockTransport(handler))
                result = await client.sync_bitable_media(target_root=Path(directory))

                self.assertEqual(result["downloaded"], [])
                self.assertEqual(result["skipped"], [])
                self.assertEqual(
                    result["missing_notes"],
                    [
                        {
                            "room": "未知房源",
                            "reason": "源多维表格存在已删除的房源笔记，飞书未返回具体记录",
                        }
                    ],
                )
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_bitable_app_token = previous_app_token
            settings.feishu_bitable_table_id = previous_table_id

    async def test_exports_sheet_as_xlsx(self) -> None:
        previous_app_id = settings.feishu_app_id
        previous_secret = settings.feishu_app_secret
        previous_sheet_token = settings.feishu_inventory_sheet_token
        try:
            settings.feishu_app_id = "cli_xxx"
            settings.feishu_app_secret = "secret_xxx"
            settings.feishu_inventory_sheet_token = "sheet_token"

            def handler(request: httpx.Request) -> httpx.Response:
                if request.url.path.endswith("/auth/v3/tenant_access_token/internal"):
                    return httpx.Response(
                        200,
                        json={"code": 0, "tenant_access_token": "tenant_token"},
                    )
                self.assertEqual(request.headers["Authorization"], "Bearer tenant_token")
                if request.url.path.endswith("/drive/v1/export_tasks"):
                    return httpx.Response(200, json={"code": 0, "data": {"ticket": "ticket_1"}})
                if request.url.path.endswith("/drive/v1/export_tasks/ticket_1"):
                    self.assertEqual(request.url.params.get("token"), "sheet_token")
                    return httpx.Response(
                        200,
                        json={
                            "code": 0,
                            "data": {
                                "result": {
                                    "job_status": "success",
                                    "file_token": "export_file",
                                }
                            },
                        },
                    )
                if request.url.path.endswith("/drive/v1/export_tasks/file/export_file/download"):
                    return httpx.Response(200, content=b"xlsx-bytes")
                raise AssertionError(f"Unexpected request: {request.url}")

            with tempfile.TemporaryDirectory() as directory:
                target_path = Path(directory) / "inventory.xlsx"
                client = FeishuClient(transport=httpx.MockTransport(handler))
                await client.export_sheet_xlsx(target_path=target_path)

                self.assertEqual(target_path.read_bytes(), b"xlsx-bytes")
        finally:
            settings.feishu_app_id = previous_app_id
            settings.feishu_app_secret = previous_secret
            settings.feishu_inventory_sheet_token = previous_sheet_token

    async def test_inventory_image_syncer_renders_only_when_sheet_changes(self) -> None:
        class FakeFeishuClient:
            def __init__(self) -> None:
                self.version = "old"
                self.exports = 0

            async def read_spreadsheet_values(self, *, spreadsheet_token: str) -> dict:
                self.exports += 1
                revision = 1 if self.version == "old" else 2
                return {
                    "sheet_id": "sheet_1",
                    "title": "总",
                    "range": "sheet_1!A1:B2",
                    "revision": revision,
                    "values": [
                        ["小区", "房号"],
                        ["星桥", self.version],
                    ],
                }

            async def export_sheet_xlsx(self, *, sheet_token: str, target_path: Path) -> Path:
                target_path.write_bytes(b"xlsx")
                return target_path

        class FakeInventoryImageSyncer(InventoryImageSyncer):
            def render_xlsx_to_inventory_images_from_values(
                self,
                xlsx_path: Path,
                values: list[list],
            ):
                return self.render_values_to_inventory_images(values)

        previous_sheet_token = settings.feishu_inventory_sheet_token
        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        previous_state_path = settings.inventory_image_sync_state_path
        previous_check_seconds = settings.feishu_inventory_sheet_check_seconds
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                settings.feishu_inventory_sheet_token = "sheet_token"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"
                settings.inventory_image_sync_state_path = Path(directory) / "state.json"
                settings.feishu_inventory_sheet_check_seconds = 0

                fake_client = FakeFeishuClient()
                syncer = FakeInventoryImageSyncer(client=fake_client)
                first_result = await syncer.refresh_if_changed()
                second_result = await syncer.refresh_if_changed()
                fake_client.version = "new"
                third_result = await syncer.refresh_if_changed()

                self.assertTrue(first_result["changed"])
                self.assertFalse(second_result["changed"])
                self.assertEqual(second_result["reason"], "unchanged")
                self.assertTrue(third_result["changed"])
                self.assertEqual(fake_client.exports, 3)
                self.assertTrue((root / "inventory_01.png").is_file())
                self.assertEqual(
                    sorted(path.name for path in root.glob("inventory_*.png")),
                    ["inventory_01.png"],
                )
        finally:
            settings.feishu_inventory_sheet_token = previous_sheet_token
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path
            settings.inventory_image_sync_state_path = previous_state_path
            settings.feishu_inventory_sheet_check_seconds = previous_check_seconds

    def test_inventory_image_render_ignores_styled_blank_cells(self) -> None:
        from openpyxl.styles import PatternFill
        from PIL import Image

        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                xlsx_path = Path(directory) / "inventory.xlsx"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"

                workbook = Workbook()
                sheet = workbook.active
                sheet["A1"] = "小区"
                sheet["B1"] = "房号"
                sheet["A2"] = "星桥"
                sheet["B2"] = "1-101"
                sheet["Z200"].fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFFF00",
                )
                workbook.save(xlsx_path)

                result = InventoryImageSyncer().render_xlsx_to_inventory_images(xlsx_path)

                self.assertEqual(result.rows, 2)
                self.assertEqual(result.columns, 2)
                with Image.open(result.paths[0]) as image:
                    width, height = image.size
                self.assertLess(width, 400)
                self.assertLess(height, 120)
        finally:
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path

    def test_inventory_image_render_draws_merged_cells_as_one_cell(self) -> None:
        from openpyxl.styles import PatternFill
        from PIL import Image

        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                xlsx_path = Path(directory) / "inventory.xlsx"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"

                workbook = Workbook()
                sheet = workbook.active
                sheet.merge_cells("A1:C1")
                sheet["A1"] = "区域标题"
                sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFD966")
                sheet["A2"] = "小区"
                sheet["B2"] = "房号"
                sheet["C2"] = "户型"
                workbook.save(xlsx_path)

                result = InventoryImageSyncer().render_xlsx_to_inventory_images(xlsx_path)

                with Image.open(result.paths[0]) as image:
                    first_column_width = InventoryImageSyncer()._column_widths(sheet, 1, 3)[0]
                    sampled = image.convert("RGB").getpixel((first_column_width + 2, 10))
                self.assertEqual(sampled, (255, 217, 102))
        finally:
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path

    def test_inventory_image_render_sanitizes_feishu_rgb_border_colors(self) -> None:
        syncer = InventoryImageSyncer()

        sanitized = syncer._sanitize_styles_xml(
            b'<styleSheet><color rgb="FFRGB(0, 0, 0)"/></styleSheet>'
        )

        self.assertIn(b'rgb="FF000000"', sanitized)

    def test_inventory_image_render_uses_inventory_header_bounds(self) -> None:
        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                xlsx_path = Path(directory) / "inventory.xlsx"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"

                workbook = Workbook()
                sheet = workbook.active
                sheet["A1"] = "标题"
                sheet["A2"] = "区域"
                sheet["B2"] = "小区"
                sheet["C2"] = "房号"
                sheet["D2"] = "户型"
                sheet["B3"] = "棠润府"
                sheet["C3"] = "1-602A"
                sheet["D3"] = "一室一厅"
                sheet["ZZ3"] = "你"
                workbook.save(xlsx_path)

                result = InventoryImageSyncer().render_xlsx_to_inventory_images(xlsx_path)

                self.assertEqual(result.columns, 4)
        finally:
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path

    def test_inventory_image_render_falls_back_without_external_renderers(self) -> None:
        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                xlsx_path = Path(directory) / "inventory.xlsx"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"

                workbook = Workbook()
                sheet = workbook.active
                sheet["A1"] = "小区"
                sheet["B1"] = "房号"
                sheet["A2"] = "星桥锦绣嘉苑"
                sheet["B2"] = "20-1606A"
                workbook.save(xlsx_path)

                with patch("app.services.inventory_image_sync.shutil.which", return_value=None):
                    result = InventoryImageSyncer().render_xlsx_to_inventory_images_from_values(
                        xlsx_path,
                        [["小区", "房号"], ["星桥锦绣嘉苑", "20-1606A"]],
                    )

                self.assertEqual(result.rows, 2)
                self.assertEqual(result.columns, 2)
                self.assertTrue(result.paths[0].is_file())
                self.assertEqual(
                    sorted(path.name for path in root.glob("inventory_*.png")),
                    ["inventory_01.png"],
                )
        finally:
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path

    def test_inventory_image_render_uses_csv_content_bounds(self) -> None:
        from PIL import Image

        previous_room_database_path = settings.room_database_path
        previous_image_glob = settings.inventory_image_glob
        previous_image_path = settings.inventory_image_path
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory) / "room_database"
                csv_path = Path(directory) / "inventory.csv"
                settings.room_database_path = root
                settings.inventory_image_glob = "room_database/inventory_*.png"
                settings.inventory_image_path = root / "inventory.png"
                csv_path.write_text(
                    "\n\n,,\n,小区,房号,,\n,星桥,1-101,,\n,,,,\n",
                    encoding="utf-8-sig",
                )

                result = InventoryImageSyncer().render_csv_to_inventory_images(csv_path)

                self.assertEqual(result.rows, 2)
                self.assertEqual(result.columns, 2)
                with Image.open(result.paths[0]) as image:
                    width, height = image.size
                self.assertLess(width, 400)
                self.assertLess(height, 120)
        finally:
            settings.room_database_path = previous_room_database_path
            settings.inventory_image_glob = previous_image_glob
            settings.inventory_image_path = previous_image_path


if __name__ == "__main__":
    unittest.main()
